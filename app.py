"""
WhatsApp → Excel Masraf Takip Sistemi
"""

import os
from dotenv import load_dotenv
load_dotenv()

import base64
import json
import logging
from datetime import datetime
from flask import Flask, request, Response
import requests
from twilio.twiml.messaging_response import MessagingResponse
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
TWILIO_AUTH_TOKEN = os.getenv("TWILIO_AUTH_TOKEN", "")
EXCEL_FILE        = os.getenv("EXCEL_FILE", "masraflar.xlsx")
USE_GOOGLE_SHEETS = os.getenv("USE_GOOGLE_SHEETS", "false").lower() == "true"
GOOGLE_SHEET_ID   = os.getenv("GOOGLE_SHEET_ID", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

app    = Flask(__name__)
client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

log.info("=== ANTHROPIC_API_KEY: %s", ANTHROPIC_API_KEY[:12] + "..." if ANTHROPIC_API_KEY else "BOŞ!")
log.info("=== TWILIO_AUTH_TOKEN: %s", TWILIO_AUTH_TOKEN[:6] + "..." if TWILIO_AUTH_TOKEN else "BOŞ!")

HEADERS = [
    "Tarih", "Saat", "Kategori", "Açıklama", "Tutar (TL)", "Para Birimi",
    "Belge Türü", "Satıcı/Kurum", "Vergi No", "KDV Tutarı",
    "Ödeme Yöntemi", "Kaynak (Numara)", "Notlar"
]

CATEGORY_COLORS = {
    "Market/Gıda":   "C8E6C9",
    "Fatura":        "BBDEFB",
    "Ulaşım":        "FFE0B2",
    "Sağlık":        "F8BBD9",
    "Eğitim":        "E1BEE7",
    "Eğlence":       "FFF9C4",
    "Giyim":         "B2EBF2",
    "Teknoloji":     "DCEDC8",
    "Restoran/Kafe": "FFCCBC",
    "Banka İşlemi":  "CFD8DC",
    "Diğer":         "F5F5F5",
}

def get_or_create_workbook():
    path = Path(EXCEL_FILE)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Masraflar"
        _write_header(ws)
        wb.save(path)
    return wb, ws

def _write_header(ws):
    header_fill  = PatternFill("solid", fgColor="1565C0")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    col_widths = [12, 8, 16, 30, 12, 12, 14, 25, 14, 12, 16, 18, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

def append_expense(data: dict, sender: str) -> int:
    wb, ws    = get_or_create_workbook()
    category  = data.get("kategori", "Diğer")
    color_hex = CATEGORY_COLORS.get(category, "F5F5F5")
    row_fill  = PatternFill("solid", fgColor=color_hex)
    now       = datetime.now()
    values    = [
        data.get("tarih") or now.strftime("%Y-%m-%d"),
        data.get("saat")  or now.strftime("%H:%M"),
        category,
        data.get("aciklama", ""),
        data.get("tutar"),
        data.get("para_birimi", "TRY"),
        data.get("belge_turu", ""),
        data.get("satici", ""),
        data.get("vergi_no", ""),
        data.get("kdv_tutari"),
        data.get("odeme_yontemi", ""),
        sender,
        data.get("notlar", ""),
    ]
    row_num = ws.max_row + 1
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = row_fill
        if col == 5:
            cell.number_format = '#,##0.00 TL'
    wb.save(EXCEL_FILE)
    log.info("Excel'e eklendi: satir %d, tutar=%s", row_num, data.get("tutar"))
    return row_num

SYSTEM_PROMPT = """Sen bir finansal belge analiz uzmanisın.
Sana gonderilen fatura, fis, banka dekontu veya benzeri belgelerdeki bilgileri
cikarin ve YALNIZCA asagidaki JSON formatinda dondur.
Tum tutarlari sayisal deger olarak ver (binslik ayrac, para birimi sembolü olmadan).
Tarih formati: YYYY-MM-DD. Bulamazsan null yaz.

{
  "tarih": "YYYY-MM-DD",
  "saat": "HH:MM",
  "kategori": "Market/Gida veya Fatura veya Ulasim veya Saglik veya Egitim veya Eglence veya Giyim veya Teknoloji veya Restoran/Kafe veya Banka Islemi veya Diger",
  "aciklama": "kisa aciklama",
  "tutar": 0.00,
  "para_birimi": "TRY",
  "belge_turu": "fis veya fatura veya dekont veya makbuz veya diger",
  "satici": "satici veya kurum adi",
  "vergi_no": null,
  "kdv_tutari": null,
  "odeme_yontemi": "nakit veya kredi karti veya banka karti veya havale veya EFT veya diger",
  "notlar": ""
}"""

def analyze_image(image_url: str, twilio_auth: tuple) -> dict | None:
    log.info("=== Goruntu indiriliyor: %s", image_url)
    log.info("=== Twilio AccountSid: %s", twilio_auth[0][:8] + "..." if twilio_auth[0] else "BOS!")
    log.info("=== Twilio AuthToken: %s",  twilio_auth[1][:6] + "..." if twilio_auth[1] else "BOS!")

    try:
        resp = requests.get(image_url, auth=twilio_auth, timeout=30)
        log.info("=== Indirme HTTP status: %d, boyut: %d bytes", resp.status_code, len(resp.content))
        resp.raise_for_status()
        mime_type = resp.headers.get("Content-Type", "image/jpeg").split(";")[0]
        log.info("=== Mime type: %s", mime_type)
        b64_data = base64.standard_b64encode(resp.content).decode("utf-8")
    except Exception as e:
        log.error("=== GORUNTU INDIRME HATASI: %s", e)
        return None

    log.info("=== Claude'a gonderiliyor...")
    try:
        message = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=1024,
            system=SYSTEM_PROMPT,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": mime_type, "data": b64_data}},
                    {"type": "text",  "text": "Bu belgeyi analiz et ve sadece JSON dondur."}
                ]
            }]
        )
        raw = message.content[0].text.strip()
        log.info("=== Claude yaniti: %s", raw[:500])
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        result = json.loads(raw.strip())
        log.info("=== JSON basariyla parse edildi!")
        return result
    except json.JSONDecodeError as e:
        log.error("=== JSON PARSE HATASI: %s", e)
        return None
    except Exception as e:
        log.error("=== CLAUDE API HATASI: %s", e)
        return None

@app.route("/webhook", methods=["POST"])
def webhook():
    sender      = request.form.get("From", "Bilinmeyen")
    num_media   = int(request.form.get("NumMedia", 0))
    body_text   = request.form.get("Body", "").strip()
    account_sid = request.form.get("AccountSid", "")
    twilio_auth = (account_sid, TWILIO_AUTH_TOKEN)

    log.info("=== Webhook: gonderen=%s, medya=%d, metin=%s", sender, num_media, body_text)

    twiml = MessagingResponse()

    if num_media == 0:
        if body_text.lower() in ("ozet", "ozet", "summary", "rapor"):
            twiml.message(_generate_summary())
        else:
            twiml.message(
                "Merhaba! Fatura, fis veya banka dekontu fotografi gonder, "
                "otomatik olarak Excel'e kaydedeyim.\n\n"
                "'ozet' yazarak son 10 masrafini gorebilirsin."
            )
        return Response(str(twiml), mimetype="application/xml")

    results = []
    for i in range(num_media):
        media_url  = request.form.get(f"MediaUrl{i}")
        media_type = request.form.get(f"MediaContentType{i}", "")
        log.info("=== Medya %d: url=%s, type=%s", i, media_url, media_type)

        if not media_url or not media_type.startswith("image/"):
            results.append("Bu dosya turunu okuyamiyorum, lutfen fotograf gonder.")
            continue

        data = analyze_image(media_url, twilio_auth)

        if data:
            row = append_expense(data, sender)
            results.append(
                f"Kaydedildi! {data.get('kategori','?')} - "
                f"{data.get('tutar','?')} {data.get('para_birimi','TRY')}\n"
                f"Satici: {data.get('satici','')}\n"
                f"Tarih: {data.get('tarih','')} | {data.get('belge_turu','')}\n"
                f"Odeme: {data.get('odeme_yontemi','')}\n"
                f"Excel satir: #{row}"
            )
        else:
            results.append("Gorsel okunamadi. Sunucu loglarini kontrol et.")

    reply = "\n\n".join(results) or "Islem tamamlandi."
    twiml.message(reply)
    return Response(str(twiml), mimetype="application/xml")

def _generate_summary() -> str:
    try:
        wb, ws = get_or_create_workbook()
        rows   = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return "Henuz kayitli masraf yok."
        last10 = rows[-10:]
        lines  = ["Son 10 Masraf:\n"]
        total  = 0
        for r in reversed(last10):
            tarih, saat, kat, acik, tutar = r[0], r[1], r[2], r[3], r[4]
            try:    total += float(tutar or 0)
            except: pass
            lines.append(f"- {tarih} | {kat} | {tutar} TRY")
        lines.append(f"\nToplam: {total:,.2f} TRY")
        return "\n".join(lines)
    except Exception as e:
        return f"Ozet alinamadi: {e}"

@app.route("/health")
def health():
    return {"status": "ok", "excel_file": EXCEL_FILE, "timestamp": datetime.now().isoformat()}

if __name__ == "__main__":
    get_or_create_workbook()
    log.info("Masraf Takip Sunucusu baslatiliyor - port 5000")
    app.run(debug=False, host="0.0.0.0", port=5000)
